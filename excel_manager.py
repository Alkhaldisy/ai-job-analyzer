from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


EXCEL_FILE = Path("job_analysis_history.xlsx")

OVERVIEW_SHEET = "Übersicht"
DETAILS_SHEET = "Job-Details"

STATUS_VALUES = [
    "Nur analysiert",
    "Interessant",
    "Bewerbung geplant",
    "Beworben",
    "Interview",
    "Absage",
    "Zusage",
]

OVERVIEW_HEADERS = [
    "Analyse-ID",
    "Analyse-Datum",
    "Unternehmen",
    "Position",
    "Standort",
    "Top Muss-Anforderungen",
    "Technologien",
    "Arbeitsmodell",
    "Arbeitszeit",
    "Bewerbungsstatus",
    "Bewerbungsdatum",
    "Stellenlink",
    "Notizen",
    "Details",
]


def create_workbook_if_missing():
    if EXCEL_FILE.exists():
        return

    workbook = Workbook()

    overview = workbook.active
    overview.title = OVERVIEW_SHEET

    details = workbook.create_sheet(DETAILS_SHEET)

    # Kopfzeile der Übersicht
    overview.append(OVERVIEW_HEADERS)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in overview[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Kopfzeile beim Scrollen sichtbar halten
    overview.freeze_panes = "A2"

    # Filter für die gesamte Übersicht
    overview.auto_filter.ref = "A1:N1"

    # Spaltenbreiten
    column_widths = {
        "A": 18,
        "B": 20,
        "C": 25,
        "D": 35,
        "E": 22,
        "F": 45,
        "G": 40,
        "H": 30,
        "I": 25,
        "J": 22,
        "K": 20,
        "L": 25,
        "M": 40,
        "N": 18,
    }

    for column, width in column_widths.items():
        overview.column_dimensions[column].width = width

    # Dropdown für Bewerbungsstatus
    status_list = ",".join(STATUS_VALUES)

    status_validation = DataValidation(
        type="list",
        formula1=f'"{status_list}"',
        allow_blank=True
    )

    overview.add_data_validation(status_validation)
    status_validation.add("J2:J1000")

    # Detail-Seite vorbereiten
    details["A1"] = "Job-Details"
    details["A1"].font = Font(
        bold=True,
        size=16
    )

    details["A2"] = (
        "Hier werden die vollständigen Details "
        "der analysierten Stellen gespeichert."
    )

    details.column_dimensions["A"].width = 28
    details.column_dimensions["B"].width = 100

    workbook.save(EXCEL_FILE)


def list_to_overview_text(values, limit=None):
    if not values:
        return ""

    if limit:
        values = values[:limit]

    return " | ".join(str(value) for value in values)


def list_to_detail_text(values):
    if not values:
        return "Keine angegeben"

    return "\n".join(
        f"• {value}"
        for value in values
    )


def generate_analysis_id(overview, analysis_datetime):
    date_prefix = analysis_datetime.strftime("%Y%m%d")

    highest_number = 0

    for row in overview.iter_rows(
        min_row=2,
        values_only=True
    ):
        existing_id = row[0]

        if not isinstance(existing_id, str):
            continue

        if not existing_id.startswith(f"{date_prefix}-"):
            continue

        try:
            number = int(existing_id.split("-")[-1])
            highest_number = max(
                highest_number,
                number
            )
        except ValueError:
            continue

    next_number = highest_number + 1

    return f"{date_prefix}-{next_number:03d}"


def save_analysis_to_excel(
    result,
    bewerbungsstatus="Nur analysiert",
    bewerbungsdatum=None,
    stellenlink="",
    notizen=""
):
    create_workbook_if_missing()

    workbook = load_workbook(EXCEL_FILE)

    overview = workbook[OVERVIEW_SHEET]
    details = workbook[DETAILS_SHEET]

    analysis_datetime = datetime.now()

    analysis_id = generate_analysis_id(
        overview,
        analysis_datetime
    )

    overview_row = overview.max_row + 1

    detail_start_row = details.max_row + 2

    top_must = list_to_overview_text(
        result.get("must_have_skills", []),
        limit=5
    )

    technologies = list_to_overview_text(
        result.get("technical_skills", []),
        limit=8
    )

    # Neue Stelle als eine Zeile in Übersicht speichern
    overview.append([
        analysis_id,
        analysis_datetime,
        result.get("company"),
        result.get("position"),
        result.get("location"),
        top_must,
        technologies,
        result.get("work_model"),
        result.get("hours"),
        bewerbungsstatus,
        bewerbungsdatum,
        "",
        notizen,
        "Details öffnen",
    ])

    # Datumsformat
    overview.cell(
        row=overview_row,
        column=2
    ).number_format = "DD.MM.YYYY HH:MM"

    if bewerbungsdatum:
        overview.cell(
            row=overview_row,
            column=11
        ).number_format = "DD.MM.YYYY"

    # Stellenlink als klickbarer Link
    link_cell = overview.cell(
        row=overview_row,
        column=12
    )

    if stellenlink:
        link_cell.value = "Stellenanzeige öffnen"
        link_cell.hyperlink = stellenlink
        link_cell.style = "Hyperlink"

    # Link zu Job-Details
    details_link_cell = overview.cell(
        row=overview_row,
        column=14
    )

    details_link_cell.value = "Details öffnen"
    details_link_cell.hyperlink = (
        f"#'{DETAILS_SHEET}'!A{detail_start_row}"
    )
    details_link_cell.style = "Hyperlink"

    # Zeilenumbruch für lange Texte
    for cell in overview[overview_row]:
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )

    # -----------------------------
    # Job-Details
    # -----------------------------

    details.merge_cells(
        start_row=detail_start_row,
        start_column=1,
        end_row=detail_start_row,
        end_column=2
    )

    title_cell = details.cell(
        row=detail_start_row,
        column=1
    )

    title_cell.value = (
        f"{analysis_id} | "
        f"{result.get('company') or 'Unbekannt'} | "
        f"{result.get('position') or 'Unbekannte Position'}"
    )

    title_cell.font = Font(
        bold=True,
        size=14,
        color="FFFFFF"
    )

    title_cell.fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    title_cell.alignment = Alignment(
        vertical="center"
    )

    # Zurück-Link
    back_row = detail_start_row + 1

    back_cell = details.cell(
        row=back_row,
        column=1
    )

    back_cell.value = "← Zurück zur Übersicht"
    back_cell.hyperlink = (
        f"#'{OVERVIEW_SHEET}'!A{overview_row}"
    )
    back_cell.style = "Hyperlink"

    detail_data = [
        (
            "Analyse-ID",
            analysis_id
        ),
        (
            "Analyse-Datum",
            analysis_datetime.strftime(
                "%d.%m.%Y %H:%M"
            )
        ),
        (
            "Unternehmen",
            result.get("company")
            or "Nicht angegeben"
        ),
        (
            "Position",
            result.get("position")
            or "Nicht angegeben"
        ),
        (
            "Standort",
            result.get("location")
            or "Nicht angegeben"
        ),
        (
            "Aufgaben",
            list_to_detail_text(
                result.get("tasks", [])
            )
        ),
        (
            "Muss-Anforderungen",
            list_to_detail_text(
                result.get(
                    "must_have_skills",
                    []
                )
            )
        ),
        (
            "Nice-to-have",
            list_to_detail_text(
                result.get(
                    "nice_to_have_skills",
                    []
                )
            )
        ),
        (
            "Technische Skills",
            list_to_detail_text(
                result.get(
                    "technical_skills",
                    []
                )
            )
        ),
        (
            "Soft Skills",
            list_to_detail_text(
                result.get(
                    "soft_skills",
                    []
                )
            )
        ),
        (
            "Sprachen",
            list_to_detail_text(
                result.get(
                    "languages",
                    []
                )
            )
        ),
        (
            "Arbeitsmodell",
            result.get("work_model")
            or "Nicht angegeben"
        ),
        (
            "Arbeitszeit",
            result.get("hours")
            or "Nicht angegeben"
        ),
        (
            "Fehlend / Unklar",
            list_to_detail_text(
                result.get(
                    "missing_or_unclear",
                    []
                )
            )
        ),
        (
            "Bewerbungsstatus",
            bewerbungsstatus
        ),
        (
            "Bewerbungsdatum",
            (
                str(bewerbungsdatum)
                if bewerbungsdatum
                else "Nicht angegeben"
            )
        ),
        (
            "Stellenlink",
            (
                stellenlink
                if stellenlink
                else "Nicht angegeben"
            )
        ),
        (
            "Notizen",
            notizen or "Keine"
        ),
    ]

    current_row = detail_start_row + 3

    for label, value in detail_data:
        label_cell = details.cell(
            row=current_row,
            column=1
        )

        value_cell = details.cell(
            row=current_row,
            column=2
        )

        label_cell.value = label
        value_cell.value = value

        label_cell.font = Font(
            bold=True
        )

        label_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        label_cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )

        value_cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )

        current_row += 1

    # Stellenlink auch in Details klickbar machen
    if stellenlink:
        link_detail_row = (
            detail_start_row
            + 3
            + 16
        )

        detail_link_cell = details.cell(
            row=link_detail_row,
            column=2
        )

        detail_link_cell.value = (
            "Stellenanzeige öffnen"
        )

        detail_link_cell.hyperlink = stellenlink
        detail_link_cell.style = "Hyperlink"

    workbook.save(EXCEL_FILE)

    return analysis_id